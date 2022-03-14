import openpyxl


def getRowCount(file, sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return (sheet.max_row)


def getColumnCount(file, sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return (sheet.max_column)


def readData(file, sheetName, rownum, columnno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.cell(row=rownum, column=columnno).value


def writeData(file, sheetName, rownum, columnno, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    sheet.cell(row=rownum, column=columnno).value = data
    workbook.save(file)


def readALLData(file, sheetName, rownum, columnno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    rownum = sheet.max_row
    columnno = sheet.max_column

    for i in range(1, rownum+1):
        for j in range(1, columnno+1):
            print(sheet.cell(i, j).value)


def storeClientData(file, sheetName, rownum, columnno):
    workbook = openpyxl.load_workbook(file)
    sheetName=workbook.active

    for i in range(1, rownum+1):
        for j in range(1, columnno+3):
            print(sheetName.cell(rownum=i, columnno=j).value)

    workbook.save(file)
